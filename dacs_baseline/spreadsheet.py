from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


CUI_NOISE = {
    "controlled unclassified information",
    "shipment identifier",
    "efts - search results",
}


@dataclass(frozen=True)
class ShipmentRow:
    identifier: str
    dd1348_expected: str | None  # Yes/No from spreadsheet if present
    identifier_type: str  # RQSTN, TCN, etc.
    source_row: int


def load_identifiers(path: Path, *, unique: bool = True) -> list[ShipmentRow]:
    """Load Shipment Identifier column from Jeff/Marianne xlsx or a .txt/.csv list."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        rows = _from_xlsx(path)
    elif suffix == ".csv":
        rows = _from_delimited(path, ",")
    else:
        rows = _from_txt(path)

    if unique:
        seen: set[str] = set()
        out: list[ShipmentRow] = []
        for row in rows:
            key = row.identifier.upper()
            if key in seen:
                continue
            seen.add(key)
            out.append(row)
        return out
    return rows


def _is_noise(value: str) -> bool:
    return value.strip().lower() in CUI_NOISE or not value.strip()


def _from_xlsx(path: Path) -> list[ShipmentRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = None
    id_col = 0
    dd_col = None
    type_col = None

    rows_data = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows_data):
        cells = [str(c).strip() if c is not None else "" for c in row]
        lowered = [c.lower() for c in cells]
        if "shipment identifier" in lowered:
            header_row = i
            id_col = lowered.index("shipment identifier")
            if "dd1348?" in lowered:
                dd_col = lowered.index("dd1348?")
            if "identifier type" in lowered:
                type_col = lowered.index("identifier type")
            break
        # fallback: first cell looks like a doc id
        if cells and re.match(r"^[A-Z0-9*]{8,}$", cells[0], re.I):
            header_row = i - 1
            break

    if header_row is None:
        raise ValueError(f"Could not find Shipment Identifier header in {path}")

    out: list[ShipmentRow] = []
    for i, row in enumerate(rows_data[header_row + 1 :], start=header_row + 2):
        cells = list(row) if row else []
        if id_col >= len(cells) or cells[id_col] is None:
            continue
        ident = str(cells[id_col]).strip()
        if _is_noise(ident):
            continue
        dd = None
        if dd_col is not None and dd_col < len(cells) and cells[dd_col] is not None:
            dd = str(cells[dd_col]).strip() or None
        itype = ""
        if type_col is not None and type_col < len(cells) and cells[type_col] is not None:
            itype = str(cells[type_col]).strip()
        out.append(ShipmentRow(ident, dd, itype, i))
    return out


def _from_txt(path: Path) -> list[ShipmentRow]:
    out: list[ShipmentRow] = []
    for i, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        ident = line.strip().strip('"')
        if not ident or ident.lower().startswith("tcn") or _is_noise(ident):
            continue
        itype = "RQSTN" if ident.endswith("*") else "TCN"
        out.append(ShipmentRow(ident, None, itype, i))
    return out


def _from_delimited(path: Path, delim: str) -> list[ShipmentRow]:
    out: list[ShipmentRow] = []
    for i, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        parts = [p.strip().strip('"') for p in line.split(delim)]
        if not parts:
            continue
        ident = parts[0]
        if not ident or (i == 1 and ident.lower() in {"tcn", "shipment identifier"}):
            continue
        if _is_noise(ident):
            continue
        itype = parts[2] if len(parts) > 2 else ("RQSTN" if ident.endswith("*") else "TCN")
        dd = parts[1] if len(parts) > 1 else None
        out.append(ShipmentRow(ident, dd, itype, i))
    return out


def infer_search_by(rows: list[ShipmentRow], override: str = "auto") -> str:
    """Return list-search radio target: tcn | document | requisition."""
    key = (override or "auto").strip().lower()
    if key in {"tcn", "document", "requisition"}:
        return key
    types = {r.identifier_type.strip().upper() for r in rows if r.identifier_type}
    # List Search radios: All | TCN | Requisition | Contract (see screenshot / EFTS map)
    if "RQSTN" in types or any(r.identifier.endswith("*") for r in rows):
        return "requisition"
    return "tcn"


def write_upload_txt(ids: list[str], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(ids) + "\n", encoding="utf-8")
    return path
